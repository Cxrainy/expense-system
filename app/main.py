"""
报销系统主应用 - Flask后端
简洁的报销申请和审批系统
"""
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_from_directory, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import os
import uuid
from werkzeug.utils import secure_filename
# 新增：导出依赖
from openpyxl import Workbook
from io import BytesIO
from sqlalchemy.orm import joinedload
# 报表导出依赖
from PIL import Image
import xlsxwriter
from tempfile import NamedTemporaryFile
import tempfile

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # 在生产环境中请使用更安全的密钥

# 文件上传配置
basedir = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(basedir, 'uploads')
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'xls', 'xlsx'}
MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB max file size

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# 确保上传目录存在
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# 数据库配置
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(basedir, "expense_system.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# 文件上传辅助函数
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_uploaded_file(file):
    """保存上传的文件并返回文件信息"""
    if file and allowed_file(file.filename):
        # 生成唯一文件名
        filename = str(uuid.uuid4()) + '_' + secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # 保存文件
        file.save(file_path)
        
        # 获取文件信息
        file_size = os.path.getsize(file_path)
        file_type = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        
        return {
            'filename': filename,  # 只存储文件名
            'original_filename': file.filename,
            'file_path': filename,  # 数据库中只存储文件名
            'file_size': file_size,
            'file_type': file_type
        }
    return None

# 数据库模型
class Notification(db.Model):
    """通知模型"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, nullable=False)
    type = db.Column(db.String(20), default='info')  # info/success/warning/error
    related_expense_id = db.Column(db.Integer, db.ForeignKey('expense.id'), nullable=True)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 关系
    user = db.relationship('User', backref=db.backref('notifications', lazy=True))
    related_expense = db.relationship('Expense', backref=db.backref('notifications', lazy=True))

class Currency(db.Model):
    """货币模型"""
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(10), unique=True, nullable=False)  # USD, CNY, EUR等
    name = db.Column(db.String(100), nullable=False)  # 美元, 人民币, 欧元等
    symbol = db.Column(db.String(10), nullable=False)  # $, ¥, €等
    exchange_rate = db.Column(db.Numeric(10, 4), nullable=False, default=1.0000)  # 相对USD的汇率
    is_active = db.Column(db.Boolean, default=True)  # 是否启用
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # 关系
    creator = db.relationship('User', backref='created_currencies')

class Category(db.Model):
    """费用分类模型"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)  # 餐饮费, 交通费等
    description = db.Column(db.Text, nullable=True)  # 分类描述
    is_active = db.Column(db.Boolean, default=True)  # 是否启用
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # 关系
    creator = db.relationship('User', backref='created_categories')

class User(db.Model):
    """用户模型"""
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), default='employee')  # employee/admin
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 关系
    expenses = db.relationship('Expense', foreign_keys='Expense.user_id', backref='submitter', lazy=True)

class ExpenseFile(db.Model):
    """报销附件模型"""
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    file_size = db.Column(db.Integer, nullable=False)
    file_type = db.Column(db.String(50), nullable=False)
    expense_id = db.Column(db.Integer, db.ForeignKey('expense.id'), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

class Expense(db.Model):
    """报销记录模型"""
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=True)
    amount = db.Column(db.Numeric(10, 2), nullable=False)  # 原币金额
    currency = db.Column(db.String(10), nullable=False, default='CNY')  # 货币类型
    exchange_rate = db.Column(db.Numeric(10, 4), nullable=False, default=1.0000)  # 汇率
    usd_amount = db.Column(db.Numeric(10, 2), nullable=False)  # 美元金额
    category = db.Column(db.String(50), nullable=False)
    expense_date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), default='pending')  # pending/approved/rejected
    
    # 审批信息
    approval_comment = db.Column(db.Text, nullable=True)
    approved_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    approved_at = db.Column(db.DateTime, nullable=True)
    
    # 关系
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # 新增关系：附件
    files = db.relationship('ExpenseFile', backref='expense', lazy=True, cascade='all, delete-orphan')

def create_notification(user_id, title, message, notification_type='info', expense_id=None):
    """创建通知"""
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        type=notification_type,
        related_expense_id=expense_id
    )
    db.session.add(notification)
    return notification

# 路由定义
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """下载上传的文件"""
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    # 记录文件访问日志
    print(f"文件访问请求: {filename}")
    print(f"完整路径: {file_path}")
    print(f"文件存在: {os.path.exists(file_path)}")
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"文件不存在: {filename}")
        # 检查数据库中是否有此文件记录
        db_file = ExpenseFile.query.filter_by(filename=filename).first()
        if db_file:
            print(f"数据库中有记录，但文件已丢失 - 文件ID: {db_file.id}, 报销ID: {db_file.expense_id}")
        else:
            print(f"数据库中无此文件记录")
        return "文件不存在", 404
    
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/expense/<int:expense_id>')
def expense_detail(expense_id):
    """报销详情页面"""
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    expense = Expense.query.get_or_404(expense_id)
    
    # 检查权限：只有申请人和管理员可以查看
    if session['role'] != 'admin' and expense.user_id != session['user_id']:
        return redirect(url_for('expenses'))
    
    return render_template('expense_detail.html', expense=expense)

@app.route('/')
def index():
    """首页 - 登录页面"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    """用户登录"""
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')
    
    user = User.query.filter_by(email=email).first()
    
    if user and user.password == password:  # 在生产环境中应该使用密码哈希
        session['user_id'] = user.id
        session['username'] = user.username
        session['role'] = user.role
        return jsonify({'success': True, 'redirect': '/dashboard'})
    else:
        return jsonify({'success': False, 'message': '邮箱或密码错误'})

@app.route('/logout')
def logout():
    """用户登出"""
    session.clear()
    return redirect(url_for('index'))

@app.route('/admin')
def admin():
    """管理页面"""
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    
    return render_template('admin.html')

# 新增：导出页面（仅管理员）
@app.route('/export')
def export_page():
    if 'user_id' not in session:
        return redirect(url_for('index'))
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    return render_template('export.html')

# 新增：报表导出页面（仅管理员）
@app.route('/reports')
def reports_page():
    if 'user_id' not in session:
        return redirect(url_for('index'))
    if session.get('role') != 'admin':
        return redirect(url_for('dashboard'))
    return render_template('reports.html')

# 尝试导入 pyecharts（不存在时不影响运行，模板将回退到原有前端渲染）
try:
    from pyecharts.charts import Line
    from pyecharts import options as opts
    from pyecharts.globals import ThemeType
    from pyecharts.commons.utils import JsCode
    from pyecharts.globals import CurrentConfig
    
    # 配置pyecharts使用本地JS文件
    CurrentConfig.ONLINE_HOST = "/static/js/pyecharts/"
    CurrentConfig.NOTEBOOK_HOST = "/static/js/pyecharts/"
    
    PYECHARTS_AVAILABLE = True
except Exception:
    PYECHARTS_AVAILABLE = False

def generate_trend_chart(user_id, role):
    """生成趋势图表HTML"""
    if not PYECHARTS_AVAILABLE:
        return None
    
    try:
        import time
        start_time = time.time()
        print(f"开始生成趋势图表 - User: {user_id}, Role: {role}")
        
        from sqlalchemy import func
        from datetime import timedelta
        
        # 近7天趋势数据
        days = 7
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=days - 1)
        dates = []
        amounts = []
        
        for i in range(days):
            current_date = start_date + timedelta(days=i)
            # 格式化日期显示（月-日）
            formatted_date = current_date.strftime('%m-%d')
            dates.append(formatted_date)
            
            # 查询当日报销总额
            base_query = db.session.query(func.sum(Expense.usd_amount)).filter(
                func.date(Expense.created_at) == current_date
            )
            if role != 'admin':
                base_query = base_query.filter_by(user_id=user_id)
            
            daily_amount = base_query.scalar() or 0
            amounts.append(float(daily_amount))
        
        # 构建简化的pyecharts折线图以提高性能
        line = Line(init_opts=opts.InitOpts(
            width='100%', 
            height='260px', 
            theme=ThemeType.LIGHT,
            renderer='canvas',  # 使用canvas渲染
            js_host="/static/js/pyecharts/",  # 本地JS文件
            animation_opts=opts.AnimationOpts(animation=False)  # 禁用动画提高性能
        ))
        
        line.add_xaxis(dates)
        line.add_yaxis(
            '美元金额',
            amounts,
            is_smooth=False,  # 禁用平滑减少计算
            symbol='circle',
            symbol_size=4,    # 减小符号大小
            linestyle_opts=opts.LineStyleOpts(width=2, color='#4CAF50'),
            label_opts=opts.LabelOpts(is_show=False),
            itemstyle_opts=opts.ItemStyleOpts(color='#4CAF50')
        )
        
        line.set_global_opts(
            title_opts=opts.TitleOpts(is_show=False),
            tooltip_opts=opts.TooltipOpts(
                trigger='axis',
                formatter='{b}: ${c}'
            ),
            xaxis_opts=opts.AxisOpts(
                type_='category', 
                boundary_gap=False
            ),
            yaxis_opts=opts.AxisOpts(
                type_='value', 
                axislabel_opts=opts.LabelOpts(formatter='${value}')
            ),
            legend_opts=opts.LegendOpts(is_show=False),
            toolbox_opts=opts.ToolboxOpts(is_show=False)
        )
        
        # 渲染图表
        chart_html = line.render_embed()
        end_time = time.time()
        print(f"趋势图表生成完成 - 耗时: {end_time - start_time:.2f}秒, HTML长度: {len(chart_html)}")
        
        return chart_html
        
    except Exception as e:
        print(f"生成趋势图表失败: {e}")
        import traceback
        traceback.print_exc()
        return None

@app.route('/dashboard')
def dashboard():
    """仪表板页面"""
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    user_id = session['user_id']
    role = session['role']
    
    # 统计数据
    if role == 'admin':
        total_expenses = Expense.query.count()
        pending_expenses = Expense.query.filter_by(status='pending').count()
        approved_expenses = Expense.query.filter_by(status='approved').count()
    else:
        total_expenses = Expense.query.filter_by(user_id=user_id).count()
        pending_expenses = Expense.query.filter_by(user_id=user_id, status='pending').count()
        approved_expenses = Expense.query.filter_by(user_id=user_id, status='approved').count()
    
    # 生成pyecharts趋势图表
    trend_chart_html = generate_trend_chart(user_id, role)
    
    return render_template(
        'dashboard.html', 
        total_expenses=total_expenses,
        pending_expenses=pending_expenses,
        approved_expenses=approved_expenses,
        trend_chart_html=trend_chart_html
    )

@app.route('/expenses')
def expenses():
    """报销申请页面"""
    if 'user_id' not in session:
        return redirect(url_for('index'))
    return render_template('expenses.html')

@app.route('/approvals')
def approvals():
    """审批管理页面 - 仅管理员"""
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    if session['role'] != 'admin':
        return redirect(url_for('expenses'))
    
    return render_template('approvals.html')

@app.route('/approval/<int:expense_id>')
def approval_detail(expense_id):
    """审批详情页面 - 仅管理员"""
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    if session['role'] != 'admin':
        return redirect(url_for('expenses'))
    
    expense = Expense.query.get_or_404(expense_id)
    
    return render_template('approval_detail.html', expense=expense)

# API 端点
@app.route('/api/expenses', methods=['GET'])
def get_expenses():
    """获取报销记录"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    role = session['role']
    status = request.args.get('status', 'all')
    currency = request.args.get('currency', 'all')  # 货币筛选
    category = request.args.get('category', 'all')  # 新增分类筛选
    limit = request.args.get('limit', type=int)  # 添加limit参数支持
    
    query = Expense.query
    
    # 普通用户只能看自己的记录
    if role != 'admin':
        query = query.filter_by(user_id=user_id)
    
    # 状态筛选
    if status and status != 'all':
        query = query.filter_by(status=status)
    
    # 货币筛选
    if currency and currency != 'all':
        query = query.filter_by(currency=currency)
    
    # 分类筛选
    if category and category != 'all':
        query = query.filter_by(category=category)
    
    # 排序并应用限制
    query = query.order_by(Expense.created_at.desc())
    if limit:
        query = query.limit(limit)
    
    expenses = query.all()
    
    result = []
    for expense in expenses:
        expense_data = {
            'id': expense.id,
            'title': expense.title,
            'description': expense.description,
            'amount': str(expense.amount),
            'currency': expense.currency,
            'exchange_rate': str(expense.exchange_rate),
            'usd_amount': str(expense.usd_amount),
            'category': expense.category,
            'expense_date': expense.expense_date.isoformat(),
            'status': expense.status,
            'submitter': expense.submitter.username,
            'created_at': expense.created_at.isoformat(),
            'approval_comment': expense.approval_comment,
            'files': [{
                'id': f.id,
                'filename': f.filename,
                'original_filename': f.original_filename,
                'file_size': f.file_size,
                'file_type': f.file_type,
                'uploaded_at': f.uploaded_at.isoformat()
            } for f in expense.files]
        }
        result.append(expense_data)
    
    return jsonify(result)

@app.route('/api/expenses', methods=['POST'])
def create_expense():
    """创建报销申请"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # 检查是否有文件上传
    if 'files' not in request.files:
        return jsonify({'error': '必须上传报销凭证'}), 400
    
    files = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': '必须上传报销凭证'}), 400
    
    # 获取表单数据
    title = request.form.get('title')
    description = request.form.get('description', '')
    amount = request.form.get('amount')
    currency = request.form.get('currency', 'CNY')
    exchange_rate = request.form.get('exchange_rate', '1.0000')
    category = request.form.get('category')
    expense_date = request.form.get('expense_date')
    
    # 验证必填字段
    if not all([title, amount, currency, exchange_rate, category, expense_date]):
        return jsonify({'error': '请填写所有必填字段'}), 400
    
    try:
        amount = float(amount)
        exchange_rate = float(exchange_rate)
        
        if amount <= 0:
            return jsonify({'error': '金额必须大于0'}), 400
        
        if exchange_rate <= 0:
            return jsonify({'error': '汇率必须大于0'}), 400
        
        # 计算美元金额（多少单位货币兑换1美元）
        usd_amount = amount / exchange_rate
        
        # 创建报销记录
        expense = Expense(
            title=title,
            description=description,
            amount=amount,
            currency=currency,
            exchange_rate=exchange_rate,
            usd_amount=usd_amount,
            category=category,
            expense_date=datetime.strptime(expense_date, '%Y-%m-%d').date(),
            user_id=session['user_id']
        )
        
        db.session.add(expense)
        db.session.flush()  # 获取expense.id
        
        # 保存上传的文件
        saved_files = []
        for file in files:
            if file.filename != '':
                file_info = save_uploaded_file(file)
                if file_info:
                    expense_file = ExpenseFile(
                        filename=file_info['filename'],
                        original_filename=file_info['original_filename'],
                        file_path=file_info['file_path'],
                        file_size=file_info['file_size'],
                        file_type=file_info['file_type'],
                        expense_id=expense.id
                    )
                    db.session.add(expense_file)
                    saved_files.append(file_info)
                else:
                    return jsonify({'error': f'文件 {file.filename} 格式不支持'}), 400
        
        if not saved_files:
            return jsonify({'error': '没有成功上传的文件'}), 400
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'id': expense.id,
            'files_count': len(saved_files)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'创建失败：{str(e)}'}), 500

@app.route('/api/expenses/<int:expense_id>/approve', methods=['POST'])
def approve_expense(expense_id):
    """审批通过"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'error': '权限不足'}), 403
    
    data = request.get_json()
    comment = data.get('comment', '')
    
    expense = Expense.query.get_or_404(expense_id)
    expense.status = 'approved'
    expense.approval_comment = comment
    expense.approved_by = session['user_id']
    expense.approved_at = datetime.utcnow()
    
    # 创建通知
    notification_title = f"报销申请已通过"
    notification_message = f"您的抧销申请「{expense.title}」已被管理员审批通过。\n审批意见：{comment if comment else '无'}"
    create_notification(
        user_id=expense.user_id,
        title=notification_title,
        message=notification_message,
        notification_type='success',
        expense_id=expense_id
    )
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/expenses/<int:expense_id>', methods=['DELETE'])
def delete_expense(expense_id):
    """删除报销申请"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    expense = Expense.query.get_or_404(expense_id)
    
    # 检查权限：只能删除自己的申请，且状态为pending
    if expense.user_id != user_id:
        return jsonify({'error': '权限不足'}), 403
    
    if expense.status != 'pending':
        return jsonify({'error': '只能删除待审批的申请'}), 400
    
    try:
        # 删除相关文件记录
        ExpenseFile.query.filter_by(expense_id=expense_id).delete()
        
        # 删除相关通知
        Notification.query.filter_by(related_expense_id=expense_id).delete()
        
        # 删除报销申请
        db.session.delete(expense)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'删除失败：{str(e)}'}), 500

@app.route('/api/expenses/<int:expense_id>', methods=['PUT'])
def update_expense(expense_id):
    """编辑报销申请（重新发起申请）"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    expense = Expense.query.get_or_404(expense_id)
    
    # 检查权限：只能编辑自己的申请
    if expense.user_id != user_id:
        return jsonify({'error': '权限不足'}), 403
    
    # 检查状态：只能编辑待审批或已拒绝的申请
    if expense.status not in ['pending', 'rejected']:
        return jsonify({'error': '只能编辑待审批或已拒绝的申请'}), 400
    
    # 检查是否有文件上传
    if 'files' not in request.files:
        return jsonify({'error': '必须上传报销凭证'}), 400
    
    files = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': '必须上传报销凭证'}), 400
    
    # 获取表单数据
    title = request.form.get('title')
    description = request.form.get('description', '')
    amount = request.form.get('amount')
    currency = request.form.get('currency', 'CNY')
    exchange_rate = request.form.get('exchange_rate', '1.0000')
    category = request.form.get('category')
    expense_date = request.form.get('expense_date')
    
    # 验证必填字段
    if not all([title, amount, currency, exchange_rate, category, expense_date]):
        return jsonify({'error': '请填写所有必填字段'}), 400
    
    try:
        amount = float(amount)
        exchange_rate = float(exchange_rate)
        
        if amount <= 0:
            return jsonify({'error': '金额必须大于0'}), 400
        
        if exchange_rate <= 0:
            return jsonify({'error': '汇率必须大于0'}), 400
        
        # 计算美元金额（多少单位货币兑换1美元）
        usd_amount = amount / exchange_rate
        
        # 删除原有文件记录
        ExpenseFile.query.filter_by(expense_id=expense_id).delete()
        
        # 更新报销记录基本信息
        expense.title = title
        expense.description = description
        expense.amount = amount
        expense.currency = currency
        expense.exchange_rate = exchange_rate
        expense.usd_amount = usd_amount
        expense.category = category
        expense.expense_date = datetime.strptime(expense_date, '%Y-%m-%d').date()
        
        # 重新发起申请：重置状态和审批信息
        expense.status = 'pending'
        expense.approval_comment = None
        expense.approved_by = None
        expense.approved_at = None
        expense.created_at = datetime.utcnow()  # 更新创建时间
        
        # 保存新的上传文件
        saved_files = []
        for file in files:
            if file.filename != '':
                file_info = save_uploaded_file(file)
                if file_info:
                    expense_file = ExpenseFile(
                        filename=file_info['filename'],
                        original_filename=file_info['original_filename'],
                        file_path=file_info['file_path'],
                        file_size=file_info['file_size'],
                        file_type=file_info['file_type'],
                        expense_id=expense_id
                    )
                    db.session.add(expense_file)
                    saved_files.append(file_info)
                else:
                    return jsonify({'error': f'文件 {file.filename} 格式不支持'}), 400
        
        if not saved_files:
            return jsonify({'error': '没有成功上传的文件'}), 400
        
        # 创建编辑通知
        create_notification(
            user_id=user_id,
            title="报销申请已重新提交",
            message=f"您已成功编辑并重新提交报销申请「{expense.title}」，请等待审批。",
            notification_type='info',
            expense_id=expense_id
        )
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'id': expense_id,
            'files_count': len(saved_files),
            'message': '报销申请已重新提交，等待审批'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'更新失败：{str(e)}'}), 500

@app.route('/api/expenses/<int:expense_id>/files')
def get_expense_files(expense_id):
    """获取报销申请的文件列表"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    expense = Expense.query.get_or_404(expense_id)
    
    # 检查权限：用户只能查看自己的申请，管理员可以查看所有
    if session['role'] != 'admin' and expense.user_id != session['user_id']:
        return jsonify({'error': '权限不足'}), 403
    
    files = ExpenseFile.query.filter_by(expense_id=expense_id).all()
    
    result = []
    for f in files:
        # 判断文件类型
        file_ext = f.original_filename.lower().split('.')[-1] if '.' in f.original_filename else ''
        
        # 支持预览的文件类型
        preview_types = {
            'image': ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp', 'svg'],
            'pdf': ['pdf'],
            'markdown': ['md', 'markdown'],
            'video': ['mp4', 'avi', 'mov', 'wmv', 'flv', 'webm', 'mkv']
        }
        
        file_category = 'other'
        can_preview = False
        
        for category, extensions in preview_types.items():
            if file_ext in extensions:
                file_category = category
                can_preview = True
                break
        
        file_data = {
            'id': f.id,
            'filename': f.filename,
            'original_filename': f.original_filename,
            'file_size': f.file_size,
            'file_type': f.file_type,
            'file_category': file_category,
            'can_preview': can_preview,
            'uploaded_at': f.uploaded_at.isoformat(),
            'download_url': f'/uploads/{f.filename}'
        }
        result.append(file_data)
    
    return jsonify(result)

@app.route('/api/expenses/<int:expense_id>/reject', methods=['POST'])
def reject_expense(expense_id):
    """审批拒绝"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'error': '权限不足'}), 403
    
    data = request.get_json()
    comment = data.get('comment', '')
    
    expense = Expense.query.get_or_404(expense_id)
    expense.status = 'rejected'
    expense.approval_comment = comment
    expense.approved_by = session['user_id']
    expense.approved_at = datetime.utcnow()
    
    # 创建通知
    notification_title = f"报销申请被拒绝"
    notification_message = f"您的抧销申请「{expense.title}」被管理员拒绝。\n拒绝原因：{comment if comment else '未提供'}"
    create_notification(
        user_id=expense.user_id,
        title=notification_title,
        message=notification_message,
        notification_type='error',
        expense_id=expense_id
    )
    
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/stats')
def get_stats():
    """获取统计数据"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    role = session['role']
    
    if role == 'admin':
        total = Expense.query.count()
        pending = Expense.query.filter_by(status='pending').count()
        approved = Expense.query.filter_by(status='approved').count()
        rejected = Expense.query.filter_by(status='rejected').count()
    else:
        total = Expense.query.filter_by(user_id=user_id).count()
        pending = Expense.query.filter_by(user_id=user_id, status='pending').count()
        approved = Expense.query.filter_by(user_id=user_id, status='approved').count()
        rejected = Expense.query.filter_by(user_id=user_id, status='rejected').count()
    
    return jsonify({
        'total': total,
        'pending': pending,
        'approved': approved,
        'rejected': rejected
    })

@app.route('/api/dashboard_stats')
def get_dashboard_stats():
    """获取仪表板详细统计数据"""
    try:
        if 'user_id' not in session:
            return jsonify({'error': '未登录', 'debug': '会话中缺少user_id'}), 401
        
        user_id = session['user_id']
        role = session.get('role', 'employee')
        
        print(f"Dashboard stats request - User ID: {user_id}, Role: {role}")  # 调试日志
        
        # 基础统计
        if role == 'admin':
            total_expenses = Expense.query.count()
            pending_expenses = Expense.query.filter_by(status='pending').count()
            approved_expenses = Expense.query.filter_by(status='approved').count()
            rejected_expenses = Expense.query.filter_by(status='rejected').count()
            
            # 总金额统计（使用美元金额）
            from sqlalchemy import func
            total_amount = db.session.query(func.sum(Expense.usd_amount)).scalar() or 0
            approved_amount = db.session.query(func.sum(Expense.usd_amount)).filter_by(status='approved').scalar() or 0
            pending_amount = db.session.query(func.sum(Expense.usd_amount)).filter_by(status='pending').scalar() or 0
            
            # 按类型统计（使用人民币金额）
            category_stats = db.session.query(
                Expense.category,
                func.count(Expense.id).label('count'),
                func.sum(Expense.usd_amount).label('amount')
            ).group_by(Expense.category).all()
            
            # 按货币类型统计
            currency_stats = db.session.query(
                Expense.currency,
                func.count(Expense.id).label('count'),
                func.sum(Expense.amount).label('original_amount'),
                func.sum(Expense.usd_amount).label('usd_amount')
            ).group_by(Expense.currency).all()
            
            # 按状态和货币统计
            status_currency_stats = db.session.query(
                Expense.status,
                Expense.currency,
                func.count(Expense.id).label('count'),
                func.sum(Expense.amount).label('original_amount'),
                func.sum(Expense.usd_amount).label('usd_amount')
            ).group_by(Expense.status, Expense.currency).all()
            
            # 今日统计
            today = datetime.utcnow().date()
            today_expenses = Expense.query.filter(
                func.date(Expense.created_at) == today
            ).count()
            today_approved = Expense.query.filter(
                func.date(Expense.approved_at) == today,
                Expense.status == 'approved'
            ).count()
            
        else:
            total_expenses = Expense.query.filter_by(user_id=user_id).count()
            pending_expenses = Expense.query.filter_by(user_id=user_id, status='pending').count()
            approved_expenses = Expense.query.filter_by(user_id=user_id, status='approved').count()
            rejected_expenses = Expense.query.filter_by(user_id=user_id, status='rejected').count()
            
            # 总金额统计（使用人民币金额）
            from sqlalchemy import func
            total_amount = db.session.query(func.sum(Expense.usd_amount)).filter_by(user_id=user_id).scalar() or 0
            approved_amount = db.session.query(func.sum(Expense.usd_amount)).filter_by(user_id=user_id, status='approved').scalar() or 0
            pending_amount = db.session.query(func.sum(Expense.usd_amount)).filter_by(user_id=user_id, status='pending').scalar() or 0
            
            # 按类型统计（使用人民币金额）
            category_stats = db.session.query(
                Expense.category,
                func.count(Expense.id).label('count'),
                func.sum(Expense.usd_amount).label('amount')
            ).filter_by(user_id=user_id).group_by(Expense.category).all()
            
            # 按货币类型统计
            currency_stats = db.session.query(
                Expense.currency,
                func.count(Expense.id).label('count'),
                func.sum(Expense.amount).label('original_amount'),
                func.sum(Expense.usd_amount).label('usd_amount')
            ).filter_by(user_id=user_id).group_by(Expense.currency).all()
            
            # 按状态和货币统计
            status_currency_stats = db.session.query(
                Expense.status,
                Expense.currency,
                func.count(Expense.id).label('count'),
                func.sum(Expense.amount).label('original_amount'),
                func.sum(Expense.usd_amount).label('usd_amount')
            ).filter_by(user_id=user_id).group_by(Expense.status, Expense.currency).all()
            
            # 今日统计
            today = datetime.utcnow().date()
            today_expenses = Expense.query.filter(
                Expense.user_id == user_id,
                func.date(Expense.created_at) == today
            ).count()
            today_approved = Expense.query.filter(
                Expense.user_id == user_id,
                func.date(Expense.approved_at) == today,
                Expense.status == 'approved'
            ).count()
        
        return jsonify({
            'basic_stats': {
                'total': total_expenses,
                'pending': pending_expenses,
                'approved': approved_expenses,
                'rejected': rejected_expenses
            },
            'amount_stats': {
                'total_amount': float(total_amount),
                'approved_amount': float(approved_amount),
                'pending_amount': float(pending_amount)
            },
            'category_stats': [{
                'category': stat.category,
                'count': stat.count,
                'amount': float(stat.amount)
            } for stat in category_stats],
            'currency_stats': [{
                'currency': stat.currency,
                'count': stat.count,
                'original_amount': float(stat.original_amount),
                'usd_amount': float(stat.usd_amount)
            } for stat in currency_stats],
            'status_currency_stats': [{
                'status': stat.status,
                'currency': stat.currency,
                'count': stat.count,
                'original_amount': float(stat.original_amount),
                'usd_amount': float(stat.usd_amount)
            } for stat in status_currency_stats],
            'today_stats': {
                'today_expenses': today_expenses,
                'today_approved': today_approved
            }
        })
    
    except Exception as e:
        print(f"Dashboard stats error: {str(e)}")  # 调试日志
        import traceback
        traceback.print_exc()
        return jsonify({'error': '获取统计数据失败', 'debug': str(e)}), 500

@app.route('/api/currency_stats')
def get_currency_stats():
    """获取货币统计数据"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    role = session['role']
    
    from sqlalchemy import func
    
    # 构建查询基础
    base_query = db.session.query(
        Expense.currency,
        Expense.status,
        func.count(Expense.id).label('count'),
        func.sum(Expense.amount).label('original_amount'),
        func.sum(Expense.usd_amount).label('usd_amount')
    )
    
    # 根据权限限制查询范围
    if role != 'admin':
        base_query = base_query.filter_by(user_id=user_id)
    
    # 按货币和状态分组统计
    currency_status_stats = base_query.group_by(Expense.currency, Expense.status).all()
    
    # 按货币分组统计（总计）
    currency_total_query = db.session.query(
        Expense.currency,
        func.count(Expense.id).label('count'),
        func.sum(Expense.amount).label('original_amount'),
        func.sum(Expense.usd_amount).label('usd_amount')
    )
    
    if role != 'admin':
        currency_total_query = currency_total_query.filter_by(user_id=user_id)
    
    currency_totals = currency_total_query.group_by(Expense.currency).all()
    
    # 格式化返回数据
    currency_summary = {}
    for stat in currency_totals:
        currency_summary[stat.currency] = {
            'currency': stat.currency,
            'total_count': stat.count,
            'total_original_amount': float(stat.original_amount),
            'total_usd_amount': float(stat.usd_amount),
            'status_breakdown': {
                'pending': {'count': 0, 'original_amount': 0.0, 'usd_amount': 0.0},
                'approved': {'count': 0, 'original_amount': 0.0, 'usd_amount': 0.0},
                'rejected': {'count': 0, 'original_amount': 0.0, 'usd_amount': 0.0}
            }
        }
    
    # 填充状态明细
    for stat in currency_status_stats:
        if stat.currency in currency_summary:
            currency_summary[stat.currency]['status_breakdown'][stat.status] = {
                'count': stat.count,
                'original_amount': float(stat.original_amount),
                'usd_amount': float(stat.usd_amount)
            }
    
    return jsonify({
        'currency_summary': list(currency_summary.values()),
        'total_currencies': len(currency_summary)
    })


@app.route('/api/notifications')
def get_notifications():
    """获取用户通知"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    unread_only = request.args.get('unread_only', 'false').lower() == 'true'
    limit = request.args.get('limit', type=int, default=10)
    
    query = Notification.query.filter_by(user_id=user_id)
    
    if unread_only:
        query = query.filter_by(is_read=False)
    
    notifications = query.order_by(Notification.created_at.desc()).limit(limit).all()
    
    result = []
    for notification in notifications:
        notification_data = {
            'id': notification.id,
            'title': notification.title,
            'message': notification.message,
            'type': notification.type,
            'is_read': notification.is_read,
            'created_at': notification.created_at.isoformat(),
            'related_expense_id': notification.related_expense_id
        }
        result.append(notification_data)
    
    return jsonify(result)

@app.route('/api/notifications/unread-count')
def get_unread_notifications_count():
    """获取未读通知数量"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    count = Notification.query.filter_by(user_id=user_id, is_read=False).count()
    
    return jsonify({'count': count})

@app.route('/api/notifications/<int:notification_id>/read', methods=['POST'])
def mark_notification_read(notification_id):
    """标记通知为已读"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    notification = Notification.query.filter_by(id=notification_id, user_id=user_id).first_or_404()
    
    notification.is_read = True
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/notifications/mark-all-read', methods=['POST'])
def mark_all_notifications_read():
    """标记所有通知为已读"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    Notification.query.filter_by(user_id=user_id, is_read=False).update({'is_read': True})
    db.session.commit()
    
    return jsonify({'success': True})

# 货币管理API
@app.route('/api/currencies')
def get_currencies():
    """获取所有活跃货币"""
    currencies = Currency.query.filter_by(is_active=True).order_by(Currency.code).all()
    result = []
    for currency in currencies:
        result.append({
            'id': currency.id,
            'code': currency.code,
            'name': currency.name,
            'symbol': currency.symbol,
            'exchange_rate': float(currency.exchange_rate)
        })
    return jsonify(result)

@app.route('/api/admin/currencies', methods=['GET'])
def admin_get_currencies():
    """管理员获取所有货币(包括禁用的)"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '没有权限'}), 403
    
    currencies = Currency.query.order_by(Currency.code).all()
    result = []
    for currency in currencies:
        result.append({
            'id': currency.id,
            'code': currency.code,
            'name': currency.name,
            'symbol': currency.symbol,
            'exchange_rate': float(currency.exchange_rate),
            'is_active': currency.is_active,
            'created_at': currency.created_at.isoformat()
        })
    return jsonify(result)

@app.route('/api/admin/currencies', methods=['POST'])
def create_currency():
    """创建新货币"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '没有权限'}), 403
    
    data = request.get_json()
    code = data.get('code', '').upper()
    name = data.get('name', '')
    symbol = data.get('symbol', '')
    exchange_rate = data.get('exchange_rate', 1.0)
    
    if not all([code, name, symbol]):
        return jsonify({'error': '请填写所有必填字段'}), 400
    
    # 检查货币代码是否已存在
    if Currency.query.filter_by(code=code).first():
        return jsonify({'error': '货币代码已存在'}), 400
    
    currency = Currency(
        code=code,
        name=name,
        symbol=symbol,
        exchange_rate=exchange_rate,
        created_by=session['user_id']
    )
    
    db.session.add(currency)
    db.session.commit()
    
    return jsonify({'success': True, 'currency_id': currency.id})

@app.route('/api/admin/currencies/<int:currency_id>', methods=['PUT'])
def update_currency(currency_id):
    """更新货币信息"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '没有权限'}), 403
    
    currency = Currency.query.get_or_404(currency_id)
    data = request.get_json()
    
    name = data.get('name', '').strip()
    symbol = data.get('symbol', '').strip()
    exchange_rate = data.get('exchange_rate', 1.0)
    
    if not all([name, symbol]):
        return jsonify({'error': '请填写所有必填字段'}), 400
    
    if exchange_rate <= 0:
        return jsonify({'error': '汇率必须大于0'}), 400
    
    # 更新字段
    currency.name = name
    currency.symbol = symbol
    currency.exchange_rate = exchange_rate
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '货币更新成功'})

@app.route('/api/admin/currencies/<int:currency_id>', methods=['DELETE'])
def delete_currency(currency_id):
    """删除货币(软删除)"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '没有权限'}), 403
    
    currency = Currency.query.get_or_404(currency_id)
    
    # 检查是否有报销记录使用该货币
    expense_count = Expense.query.filter_by(currency=currency.code).count()
    if expense_count > 0:
        currency.is_active = False
        db.session.commit()
        return jsonify({'success': True, 'message': '货币已禁用，因为有报销记录使用该货币'})
    else:
        db.session.delete(currency)
        db.session.commit()
        return jsonify({'success': True, 'message': '货币已删除'})

@app.route('/api/admin/currencies/<int:currency_id>/restore', methods=['POST'])
def restore_currency(currency_id):
    """恢复货币"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '没有权限'}), 403
    
    currency = Currency.query.get_or_404(currency_id)
    
    # 检查是否已被禁用
    if currency.is_active:
        return jsonify({'error': '货币已经是启用状态'}), 400
    
    currency.is_active = True
    db.session.commit()
    
    return jsonify({'success': True, 'message': '货币已恢复'})

# 分类管理API
@app.route('/api/categories')
def get_categories():
    """获取所有活跃分类"""
    categories = Category.query.filter_by(is_active=True).order_by(Category.name).all()
    result = []
    for category in categories:
        result.append({
            'id': category.id,
            'name': category.name,
            'description': category.description
        })
    return jsonify(result)

@app.route('/api/admin/categories', methods=['GET'])
def admin_get_categories():
    """管理员获取所有分类(包括禁用的)"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '没有权限'}), 403
    
    categories = Category.query.order_by(Category.name).all()
    result = []
    for category in categories:
        result.append({
            'id': category.id,
            'name': category.name,
            'description': category.description,
            'is_active': category.is_active,
            'created_at': category.created_at.isoformat()
        })
    return jsonify(result)

@app.route('/api/admin/categories', methods=['POST'])
def create_category():
    """创建新分类"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '没有权限'}), 403
    
    data = request.get_json()
    name = data.get('name', '').strip()
    description = data.get('description', '').strip()
    
    if not name:
        return jsonify({'error': '请填写分类名称'}), 400
    
    # 检查分类名称是否已存在
    if Category.query.filter_by(name=name).first():
        return jsonify({'error': '分类名称已存在'}), 400
    
    category = Category(
        name=name,
        description=description,
        created_by=session['user_id']
    )
    
    db.session.add(category)
    db.session.commit()
    
    return jsonify({'success': True, 'category_id': category.id})

@app.route('/api/admin/categories/<int:category_id>', methods=['PUT'])
def update_category(category_id):
    """更新分类信息"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '没有权限'}), 403
    
    category = Category.query.get_or_404(category_id)
    data = request.get_json()
    
    name = data.get('name', '').strip()
    description = data.get('description', '').strip()
    
    if not name:
        return jsonify({'error': '请填写分类名称'}), 400
    
    # 检查名称是否与其他分类冲突
    existing_category = Category.query.filter_by(name=name).first()
    if existing_category and existing_category.id != category_id:
        return jsonify({'error': '分类名称已存在'}), 400
    
    # 更新字段
    category.name = name
    category.description = description
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '分类更新成功'})

@app.route('/api/admin/categories/<int:category_id>', methods=['DELETE'])
def delete_category(category_id):
    """删除分类(软删除)"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '没有权限'}), 403
    
    category = Category.query.get_or_404(category_id)
    
    # 检查是否有报销记录使用该分类
    expense_count = Expense.query.filter_by(category=category.name).count()
    if expense_count > 0:
        category.is_active = False
        db.session.commit()
        return jsonify({'success': True, 'message': '分类已禁用，因为有报销记录使用该分类'})
    else:
        db.session.delete(category)
        db.session.commit()
        return jsonify({'success': True, 'message': '分类已删除'})

@app.route('/api/admin/categories/<int:category_id>/restore', methods=['POST'])
def restore_category(category_id):
    """恢复分类"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '没有权限'}), 403
    
    category = Category.query.get_or_404(category_id)
    
    # 检查是否已被禁用
    if category.is_active:
        return jsonify({'error': '分类已经是启用状态'}), 400
    
    category.is_active = True
    db.session.commit()
    
    return jsonify({'success': True, 'message': '分类已恢复'})

# 导出API：按时间区间导出报销数据为Excel（仅管理员）
@app.route('/api/export', methods=['POST'])
def export_expenses():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if session.get('role') != 'admin':
        return jsonify({'error': '无权限'}), 403

    data = request.get_json(silent=True) or {}
    start_date_str = data.get('start_date')
    end_date_str = data.get('end_date')

    # 解析日期
    start_date = None
    end_date = None
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'error': '日期格式不正确，应为YYYY-MM-DD'}), 400

    # 构建查询
    query = Expense.query.options(joinedload(Expense.submitter))
    if start_date:
        query = query.filter(Expense.expense_date >= start_date)
    if end_date:
        query = query.filter(Expense.expense_date <= end_date)

    expenses = query.order_by(Expense.expense_date.asc(), Expense.id.asc()).all()

    # 组装导出数据
    records = []
    for e in expenses:
        submitter_name = e.submitter.username if getattr(e, 'submitter', None) else ''
        submitter_email = e.submitter.email if getattr(e, 'submitter', None) else ''
        records.append({
            'ID': e.id,
            '标题': e.title,
            '描述': e.description or '',
            '分类': e.category,
            '原币金额': float(e.amount),
            '货币': e.currency,
            '汇率(单位=1美元)': float(e.exchange_rate),
            '美元金额': float(e.usd_amount),
            '报销日期': e.expense_date.strftime('%Y-%m-%d') if e.expense_date else '',
            '状态': e.status,
            '申请人': submitter_name,
            '申请人邮箱': submitter_email,
            '创建时间': e.created_at.strftime('%Y-%m-%d %H:%M:%S') if e.created_at else '',
            '更新时间': e.updated_at.strftime('%Y-%m-%d %H:%M:%S') if e.updated_at else ''
        })

    # 生成Excel (使用openpyxl替代pandas)
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = '报销数据'
    
    # 写入表头
    if records:
        headers = list(records[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 写入数据
        for row, record in enumerate(records, 2):
            for col, value in enumerate(record.values(), 1):
                ws.cell(row=row, column=col, value=value)
    
    wb.save(output)
    output.seek(0)

    # 生成文件名
    if start_date and end_date:
        filename = f"expense_export_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    elif start_date and not end_date:
        filename = f"expense_export_from_{start_date.strftime('%Y%m%d')}.xlsx"
    elif end_date and not start_date:
        filename = f"expense_export_to_{end_date.strftime('%Y%m%d')}.xlsx"
    else:
        filename = 'expense_export_all.xlsx'

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# 报表相关API
def process_image_for_excel(image_path, quality='medium'):
    """处理图片用于Excel嵌入，返回处理后的图片数据和尺寸"""
    try:
        if not os.path.exists(image_path):
            return None, None, None
        
        # 打开并处理图片
        with Image.open(image_path) as img:
            # 转换为RGB模式（如果是RGBA等）
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # 获取原始尺寸
            original_width, original_height = img.size
            aspect_ratio = original_width / original_height
            
            # 根据质量设置计算目标尺寸 - 更合理的尺寸设置
            if quality == 'high':
                # 高质量：适中尺寸，保证清晰度
                target_width, target_height = 300, 200
            elif quality == 'low':
                # 低质量：小尺寸，节省空间
                target_width, target_height = 150, 100
            else:  # medium
                # 中等质量：平衡尺寸和质量
                target_width, target_height = 200, 150
            
            # 智能计算最终尺寸，保持宽高比
            if aspect_ratio > 1.5:  # 宽图片（如横向照片）
                final_width = target_width
                final_height = target_width / aspect_ratio
                # 限制最大高度
                if final_height > target_height:
                    final_height = target_height
                    final_width = target_height * aspect_ratio
            elif aspect_ratio < 0.75:  # 高图片（如纵向照片）
                final_height = target_height
                final_width = target_height * aspect_ratio
                # 限制最大宽度
                if final_width > target_width:
                    final_width = target_width
                    final_height = target_width / aspect_ratio
            else:  # 接近正方形的图片
                # 使用较小的目标尺寸以保持比例
                min_dimension = min(target_width, target_height)
                if aspect_ratio >= 1:
                    final_width = min_dimension
                    final_height = min_dimension / aspect_ratio
                else:
                    final_height = min_dimension
                    final_width = min_dimension * aspect_ratio
            
            # 确保尺寸不会太小
            final_width = max(final_width, 80)
            final_height = max(final_height, 60)
            
            # 调整图片尺寸 - 使用高质量重采样
            target_size = (int(final_width * 2), int(final_height * 2))  # 先放大2倍
            img = img.resize(target_size, Image.Resampling.LANCZOS)
            
            # 再缩小到最终尺寸，提高图片质量
            final_size = (int(final_width), int(final_height))
            img = img.resize(final_size, Image.Resampling.LANCZOS)
            
            # 保存到内存
            img_buffer = BytesIO()
            save_quality = 95 if quality == 'high' else (85 if quality == 'medium' else 70)
            img.save(img_buffer, format='JPEG', quality=save_quality, optimize=True)
            img_buffer.seek(0)
            
            return img_buffer, final_width, final_height
            
    except Exception as e:
        print(f"图片处理失败: {e}")
        return None, None, None

@app.route('/api/reports/preview', methods=['POST'])
def preview_report_data():
    """预览报表数据"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '权限不足'}), 403
    
    try:
        data = request.get_json()
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        status_filter = data.get('status', 'all')
        
        # 解析日期
        start_date = None
        end_date = None
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # 构建查询
        query = Expense.query.options(joinedload(Expense.submitter), joinedload(Expense.files))
        
        if start_date:
            query = query.filter(Expense.expense_date >= start_date)
        if end_date:
            query = query.filter(Expense.expense_date <= end_date)
        if status_filter and status_filter != 'all':
            query = query.filter(Expense.status == status_filter)
        
        expenses = query.all()
        
        # 统计数据
        total_records = len(expenses)
        total_amount = sum(float(e.usd_amount) for e in expenses)
        with_images = len([e for e in expenses if any(f.file_type.lower() in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp'] for f in e.files)])
        
        # 按类型分组统计
        category_stats = {}
        for expense in expenses:
            category = expense.category
            if category not in category_stats:
                category_stats[category] = {'count': 0, 'amount': 0}
            category_stats[category]['count'] += 1
            category_stats[category]['amount'] += float(expense.usd_amount)
        
        category_breakdown = [
            {
                'category': cat,
                'count': stats['count'],
                'amount': stats['amount']
            }
            for cat, stats in category_stats.items()
        ]
        category_breakdown.sort(key=lambda x: x['amount'], reverse=True)
        
        return jsonify({
            'success': True,
            'data': {
                'total_records': total_records,
                'total_categories': len(category_stats),
                'total_amount': total_amount,
                'with_images': with_images,
                'category_breakdown': category_breakdown
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'预览失败：{str(e)}'}), 500

@app.route('/api/reports/export-summary', methods=['POST'])
def export_monthly_summary():
    """生成月度汇总报表"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '需要管理员权限'}), 403
    
    try:
        # 支持JSON和表单两种数据格式
        if request.content_type and 'application/json' in request.content_type:
            data = request.get_json()
        else:
            # 表单数据
            data = request.form.to_dict()
        
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # 构建查询
        query = Expense.query.filter(Expense.status == 'approved')
        
        if start_date:
            query = query.filter(Expense.expense_date >= start_date)
        if end_date:
            query = query.filter(Expense.expense_date <= end_date)
        
        expenses = query.all()
        
        if not expenses:
            return jsonify({'error': '所选时间范围内没有已批准的报销记录'}), 400
        
        # 生成月度汇总Excel
        filename = generate_monthly_summary_excel(expenses, start_date, end_date)
        
        return send_file(
            filename,
            as_attachment=True,
            download_name=os.path.basename(filename),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'导出失败：{str(e)}'}), 500

@app.route('/api/reports/export', methods=['POST'])
def export_detailed_report():
    """导出详细报表（包含图片）"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '权限不足'}), 403
    
    try:
        # 获取表单数据
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        status_filter = request.form.get('status', 'all')
        include_images = request.form.get('include_images') == 'true'
        group_by_category = request.form.get('group_by_category') == 'true'
        include_comments = request.form.get('include_comments') == 'true'
        image_quality = request.form.get('image_quality', 'medium')
        
        # 解析日期
        start_date = None
        end_date = None
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # 构建查询
        query = Expense.query.options(joinedload(Expense.submitter), joinedload(Expense.files))
        
        if start_date:
            query = query.filter(Expense.expense_date >= start_date)
        if end_date:
            query = query.filter(Expense.expense_date <= end_date)
        if status_filter and status_filter != 'all':
            query = query.filter(Expense.status == status_filter)
        
        expenses = query.order_by(Expense.expense_date.asc(), Expense.category, Expense.id).all()
        
        # 创建临时文件
        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            workbook = xlsxwriter.Workbook(tmp_file.name, {'strings_to_numbers': True})
            
            # 定义格式
            header_format = workbook.add_format({
                'bold': True,
                'font_size': 12,
                'bg_color': '#366092',
                'font_color': 'white',
                'border': 2,
                'align': 'center',
                'valign': 'vcenter',
                'font_name': 'Microsoft YaHei'
            })
            
            cell_format = workbook.add_format({
                'border': 1,
                'align': 'left',
                'valign': 'vcenter',
                'text_wrap': True,
                'font_name': 'Microsoft YaHei',
                'font_size': 10
            })
            
            cell_center_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': True,
                'font_name': 'Microsoft YaHei',
                'font_size': 10
            })
            
            # 添加交替行格式（浅灰色背景）
            cell_center_alt_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': True,
                'font_name': 'Microsoft YaHei',
                'font_size': 10,
                'bg_color': '#F8F9FA'
            })
            
            cell_alt_format = workbook.add_format({
                'border': 1,
                'align': 'left',
                'valign': 'vcenter',
                'text_wrap': True,
                'font_name': 'Microsoft YaHei',
                'font_size': 10,
                'bg_color': '#F8F9FA'
            })
            
            amount_format = workbook.add_format({
                'border': 1,
                'align': 'right',
                'valign': 'vcenter',
                'num_format': '#,##0.00',
                'font_name': 'Microsoft YaHei',
                'font_size': 10
            })
            
            usd_amount_format = workbook.add_format({
                'border': 1,
                'align': 'right',
                'valign': 'vcenter',
                'num_format': '$#,##0.00',
                'font_name': 'Microsoft YaHei',
                'font_size': 10,
                'bg_color': '#E3F2FD',  # 更明显的蓝色背景
                'bold': True  # 加粗美元金额
            })
            
            date_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'num_format': 'yyyy-mm-dd',
                'font_name': 'Microsoft YaHei',
                'font_size': 10
            })
            
            currency_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'font_name': 'Microsoft YaHei',
                'font_size': 10,
                'bold': True,
                'bg_color': '#E8F5E8',  # 淡绿色背景，更醒目
                'font_color': '#2E7D32'  # 深绿色字体
            })
            
            if group_by_category:
                # 按类型分工作簿
                expenses_by_category = {}
                for expense in expenses:
                    category = expense.category
                    if category not in expenses_by_category:
                        expenses_by_category[category] = []
                    expenses_by_category[category].append(expense)
                
                for category, category_expenses in expenses_by_category.items():
                    create_worksheet(workbook, category, category_expenses, include_images, 
                                   include_comments, image_quality, header_format, cell_format, 
                                   cell_center_format, cell_alt_format, cell_center_alt_format,
                                   amount_format, usd_amount_format, date_format, currency_format)
            else:
                # 单个工作簿
                create_worksheet(workbook, '全部报销', expenses, include_images, 
                               include_comments, image_quality, header_format, cell_format, 
                               cell_center_format, cell_alt_format, cell_center_alt_format,
                               amount_format, usd_amount_format, date_format, currency_format)
            
            workbook.close()
            
            # 生成文件名
            if start_date and end_date:
                filename = f"detailed_report_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
            elif start_date:
                filename = f"detailed_report_from_{start_date.strftime('%Y%m%d')}.xlsx"
            elif end_date:
                filename = f"detailed_report_to_{end_date.strftime('%Y%m%d')}.xlsx"
            else:
                filename = 'detailed_report_all.xlsx'
            
            return send_file(
                tmp_file.name,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    
    except Exception as e:
        return jsonify({'error': f'导出失败：{str(e)}'}), 500

def generate_monthly_summary_excel(expenses, start_date, end_date):
    """生成月度汇总Excel报表"""
    # 创建临时文件
    temp_dir = tempfile.mkdtemp()
    
    # 生成文件名
    date_suffix = f"{start_date}_to_{end_date}" if start_date and end_date else "all_time"
    filename = os.path.join(temp_dir, f'月度汇总报表_{date_suffix}.xlsx')
    
    # 创建Excel工作簿
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet('月度汇总')
    
    # 定义预设的报销类型和标准占比
    category_mapping = {
        '车船费': {'序号': 1, '标准占比': 0.01},
        '过桥费停车费': {'序号': 1, '标准占比': 0.01},
        '维修费': {'序号': 2, '标准占比': 0.01},
        '办公费': {'序号': 3, '标准占比': 0.01},
        '油费': {'序号': 4, '标准占比': 0.01},
        '生活补助': {'序号': 5, '标准占比': 0.01},
        '水电费': {'序号': 5, '标准占比': 0.01},
        '住宿费': {'序号': 6, '标准占比': 0.01},
        '福利费': {'序号': 7, '标准占比': 0.01},
        '招待费': {'序号': 8, '标准占比': 0.01},
        '广告费': {'序号': 10, '标准占比': 0.01},
        '运输费': {'序号': 9, '标准占比': 0.05},
        '手续费': {'序号': 13, '标准占比': 0.01},
        '营业外支出': {'序号': 10, '标准占比': 0.01},
        '工资': {'序号': 11, '标准占比': 0.05},
        '租赁费': {'序号': 12, '标准占比': 0.01},
        '其他应收': {'序号': 13, '标准占比': 0.00},
    }
    
    # 按报销内容分组统计
    category_stats = {}
    total_usd_amount = 0
    
    for expense in expenses:
        try:
            category = expense.category or '其他'
            usd_amount = float(expense.usd_amount or 0)
            total_usd_amount += usd_amount
            
            if category not in category_stats:
                category_stats[category] = {
                    'total_usd': 0,
                    'total_cny': 0,
                    'remarks': []
                }
            
            category_stats[category]['total_usd'] += usd_amount
            category_stats[category]['total_cny'] += usd_amount * 7.12  # 固定汇率7.12
            
            # 构建备注信息：备注内容+金额/汇率
            remark_text = expense.description or ''
            original_amount = float(expense.amount or 0)
            # 根据原币种计算汇率
            if expense.currency == 'USD':
                # 如果原币种是美元，汇率为1
                exchange_rate = 1
            else:
                # 计算原始汇率 (原币金额/美元金额)
                exchange_rate = float(expense.exchange_rate) if expense.exchange_rate else (original_amount / usd_amount if usd_amount > 0 else 0)
            
            # 格式：备注内容+金额/汇率
            if remark_text.strip():
                remark_entry = f"{remark_text.strip()}支出{original_amount:.2f}/汇率{exchange_rate:.0f}"
            else:
                remark_entry = f"{original_amount:.2f}/汇率{exchange_rate:.0f}"
            category_stats[category]['remarks'].append(remark_entry)
            
        except Exception as e:
            print(f"处理报销记录时出错: {e}, 记录ID: {expense.id if hasattr(expense, 'id') else 'unknown'}")
            continue
    
    # 设置Excel样式
    header_format = workbook.add_format({
        'bold': True,
        'font_size': 12,
        'bg_color': '#366092',
        'font_color': 'white',
        'border': 2,
        'align': 'center',
        'valign': 'vcenter',
        'font_name': 'Microsoft YaHei'
    })
    
    cell_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'font_name': 'Microsoft YaHei',
        'font_size': 10
    })
    
    amount_format = workbook.add_format({
        'border': 1,
        'align': 'right',
        'valign': 'vcenter',
        'num_format': '#,##0.00',
        'font_name': 'Microsoft YaHei',
        'font_size': 10
    })
    
    percentage_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'num_format': '0.00%',
        'font_name': 'Microsoft YaHei',
        'font_size': 10
    })
    
    # 设置列宽
    worksheet.set_column('A:A', 8)   # 序号
    worksheet.set_column('B:B', 15)  # 报销内容
    worksheet.set_column('C:C', 15)  # 报销金额/美元
    worksheet.set_column('D:D', 15)  # 报销金额/人民币
    worksheet.set_column('E:E', 12)  # 标准占比
    worksheet.set_column('F:F', 12)  # 实际占比
    worksheet.set_column('G:G', 50)  # 备注
    
    # 写入表头
    headers = ['序号', '报销内容', '报销金额/美元', '报销金额/人民币', '标准占比', '实际占比', '备注']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    # 写入数据
    row = 1
    for category, mapping in category_mapping.items():
        if category in category_stats:
            stats = category_stats[category]
            usd_amount = stats['total_usd']
            cny_amount = stats['total_cny']
            actual_ratio = usd_amount / total_usd_amount if total_usd_amount > 0 else 0
            remarks = '，'.join(stats['remarks'])
            
            worksheet.write(row, 0, mapping['序号'], cell_format)
            worksheet.write(row, 1, category, cell_format)
            worksheet.write(row, 2, usd_amount, amount_format)
            worksheet.write(row, 3, cny_amount, amount_format)
            worksheet.write(row, 4, mapping['标准占比'], percentage_format)
            worksheet.write(row, 5, actual_ratio, percentage_format)
            worksheet.write(row, 6, remarks, cell_format)
        else:
            # 没有数据的类型显示为NaN
            worksheet.write(row, 0, mapping['序号'], cell_format)
            worksheet.write(row, 1, category, cell_format)
            worksheet.write(row, 2, 'NaN', cell_format)
            worksheet.write(row, 3, 'NaN', cell_format)
            worksheet.write(row, 4, mapping['标准占比'], percentage_format)
            worksheet.write(row, 5, '手动统计', cell_format)
            worksheet.write(row, 6, 'NaN', cell_format)
        
        row += 1
    
    workbook.close()
    return filename

def create_worksheet(workbook, sheet_name, expenses, include_images, include_comments, 
                    image_quality, header_format, cell_format, cell_center_format,
                    cell_alt_format, cell_center_alt_format, amount_format, 
                    usd_amount_format, date_format, currency_format):
    """创建工作表"""
    # 确保工作表名称有效
    safe_sheet_name = sheet_name.replace('/', '_').replace('\\', '_')[:31]  # Excel工作表名限制31字符
    worksheet = workbook.add_worksheet(safe_sheet_name)
    
    # 设置列宽 - 优化图片列宽度
    worksheet.set_column('A:A', 12)  # 报销日期
    worksheet.set_column('B:B', 15)  # 报销类型
    worksheet.set_column('C:C', 25)  # 报销标题
    worksheet.set_column('D:D', 10)  # 原币种
    worksheet.set_column('E:E', 15)  # 原币金额
    worksheet.set_column('F:F', 15)  # 报销金额(USD)
    worksheet.set_column('G:G', 25)  # 报销备注（稍微缩小给图片让空间）
    if include_comments:
        worksheet.set_column('H:H', 18)  # 审批意见
        worksheet.set_column('I:I', 12)  # 申请人
        if include_images:
            worksheet.set_column('J:J', 30)  # 图片/文件（增加宽度）
    else:
        worksheet.set_column('H:H', 12)  # 申请人
        if include_images:
            worksheet.set_column('I:I', 30)  # 图片/文件（增加宽度）
    
    # 写入表头
    headers = ['报销日期', '报销类型', '报销标题', '原币种', '原币金额', '报销金额(USD)', '报销备注']
    if include_comments:
        headers.append('审批意见')
    headers.append('申请人')
    if include_images:
        headers.append('附件')
    
    # 添加工作表标题（第一行）
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 16,
        'align': 'center',
        'valign': 'vcenter',
        'font_name': 'Microsoft YaHei',
        'bg_color': '#2E5984',
        'font_color': 'white',
        'border': 2
    })
    
    # 合并单元格作为标题
    worksheet.merge_range(0, 0, 0, len(headers)-1, f'{sheet_name} - 报销记录明细表', title_format)
    worksheet.set_row(0, 25)  # 设置标题行高
    
    # 写入表头（第二行）
    for col, header in enumerate(headers):
        worksheet.write(1, col, header, header_format)
    worksheet.set_row(1, 18)  # 设置表头行高
    
    # 写入数据
    current_row = 2  # 从第三行开始（0是标题，1是表头）
    for idx, expense in enumerate(expenses):
        row_height = 20  # 默认行高
        
        # 确定是否使用交替行格式
        is_alternate_row = idx % 2 == 1
        row_cell_format = cell_alt_format if is_alternate_row else cell_format
        row_center_format = cell_center_alt_format if is_alternate_row else cell_center_format
        
        # 基本数据
        worksheet.write(current_row, 0, expense.expense_date, date_format)
        worksheet.write(current_row, 1, expense.category, row_center_format)
        worksheet.write(current_row, 2, expense.title, row_cell_format)
        worksheet.write(current_row, 3, expense.currency, currency_format)
        worksheet.write(current_row, 4, float(expense.amount), amount_format)
        worksheet.write(current_row, 5, float(expense.usd_amount), usd_amount_format)
        worksheet.write(current_row, 6, expense.description or '', row_cell_format)
        
        col_offset = 7
        if include_comments:
            worksheet.write(current_row, col_offset, expense.approval_comment or '', cell_format)
            col_offset += 1
        
        # 申请人
        submitter_name = expense.submitter.username if hasattr(expense, 'submitter') and expense.submitter else ''
        worksheet.write(current_row, col_offset, submitter_name, row_center_format)
        col_offset += 1
        
        # 处理附件
        if include_images and expense.files:
            file_info = []
            images_added = 0
            
            for file_obj in expense.files:
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], file_obj.filename)
                file_ext = file_obj.file_type.lower()
                
                if file_ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp']:
                    # 处理图片
                    img_data, img_width, img_height = process_image_for_excel(file_path, image_quality)
                    if img_data and images_added < 3:  # 限制每行最多3张图片
                        try:
                            # Excel列宽设置 - 附件列较宽
                            cell_width_pixels = 210  # 30*7像素，约210像素
                            min_cell_height = 90     # 最小单元格高度，给图片更多空间
                            
                            # 计算图片在单元格中的最佳尺寸和位置
                            available_width = cell_width_pixels - 8  # 留出边距
                            available_height = max(min_cell_height, img_height + 10) - 8
                            
                            # 计算缩放比例，保持宽高比
                            scale_x = available_width / img_width if img_width > available_width else 1.0
                            scale_y = available_height / img_height if img_height > available_height else 1.0
                            scale = min(scale_x, scale_y, 1.0)  # 只缩小，不放大
                            
                            # 最终显示尺寸
                            display_width = img_width * scale
                            display_height = img_height * scale
                            
                            # 居中偏移计算
                            x_offset = max(4, (cell_width_pixels - display_width) / 2)
                            y_offset = max(4, (available_height + 8 - display_height) / 2)
                            
                            # 插入图片
                            worksheet.insert_image(current_row, col_offset, file_path, {
                                'image_data': img_data,
                                'x_scale': scale,
                                'y_scale': scale,
                                'x_offset': int(x_offset),
                                'y_offset': int(y_offset),
                                'positioning': 1  # 移动但不调整大小
                            })
                            images_added += 1
                            
                            # 调整行高以完全容纳图片
                            required_height = display_height + 12  # 图片高度 + 上下边距
                            row_height = max(row_height, required_height)
                            
                            file_info.append(f"图片: {file_obj.original_filename}")
                        except Exception as e:
                            file_info.append(f"图片加载失败: {file_obj.original_filename}")
                    else:
                        file_info.append(f"图片: {file_obj.original_filename}")
                else:
                    # 其他文件
                    file_info.append(f"文件: {file_obj.original_filename} (位置: {file_path})")
            
            # 写入文件信息
            worksheet.write(current_row, col_offset, '\n'.join(file_info), cell_format)
        
        # 设置行高 - 确保最小行高以获得更好的视觉效果
        final_row_height = max(row_height, 25)  # 最少25像素高度
        worksheet.set_row(current_row, final_row_height)
        current_row += 1
    
    # 添加统计信息
    if expenses:
        current_row += 2  # 空一行
        
        # 创建统计格式
        summary_format = workbook.add_format({
            'bold': True,
            'font_size': 11,
            'bg_color': '#FFE0B2',  # 温暖的橙色背景
            'border': 2,
            'align': 'center',
            'valign': 'vcenter',
            'font_name': 'Microsoft YaHei',
            'font_color': '#E65100'  # 深橙色字体
        })
        
        # 创建金额统计格式
        summary_amount_format = workbook.add_format({
            'bold': True,
            'font_size': 11,
            'bg_color': '#FFF3E0',  # 浅橙色背景
            'border': 2,
            'align': 'right',
            'valign': 'vcenter',
            'font_name': 'Microsoft YaHei',
            'num_format': '#,##0.00'
        })
        
        summary_usd_format = workbook.add_format({
            'bold': True,
            'font_size': 11,
            'bg_color': '#E1F5FE',  # 浅蓝色背景
            'border': 2,
            'align': 'right',
            'valign': 'vcenter',
            'font_name': 'Microsoft YaHei',
            'num_format': '$#,##0.00'
        })
        
        # 统计信息
        total_original_amount = sum(float(e.amount) for e in expenses)
        total_usd_amount = sum(float(e.usd_amount) for e in expenses)
        
        worksheet.write(current_row, 3, '原币总计:', summary_format)
        worksheet.write(current_row, 4, total_original_amount, summary_amount_format)
        worksheet.write(current_row, 5, '美元总计:', summary_format)
        worksheet.write(current_row, 6, total_usd_amount, summary_usd_format)
        
        # 记录统计
        current_row += 1
        worksheet.merge_range(current_row, 3, current_row, 6, f'共计记录: {len(expenses)}条', summary_format)
        
        # 设置统计行的行高
        worksheet.set_row(current_row - 1, 22)
        worksheet.set_row(current_row, 22)

# 报销类型导入功能
@app.route('/api/admin/import-categories', methods=['POST'])
def import_expense_categories():
    """导入预定义的报销类型"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '需要管理员权限'}), 403
    
    try:
        # 预定义的报销类型数据
        categories_data = [
            ("过桥费停车费", "车辆通行过桥费、隧道费、停车费等交通相关费用。包括高速公路过路费、市内停车场费用、临时停车费等。"),
            ("维修费", "设备、车辆、办公设施等的维护和修理费用。包括预防性维护、故障修复、零部件更换、保养服务等。"),
            ("办公费", "日常办公用品和办公设备的采购费用。包括文具用品、纸张、墨盒、办公设备小件、办公软件等。"),
            ("油费", "车辆燃油费用，包括汽油、柴油等。用于公务用车、出差车辆的加油费用报销。"),
            ("生活补助", "员工生活相关的补助费用。包括餐补、交通补助、通讯补助、节日补助等福利性支出。"),
            ("水电费", "办公场所的水费、电费、燃气费等基础设施费用。包括月度账单、临时用水用电费用。"),
            ("住宿费", "出差期间的酒店住宿费用。包括标准间、商务间等符合公司差旅标准的住宿开支。"),
            ("福利费", "员工福利相关支出。包括节日慰问、生日福利、团建活动、员工体检、培训等福利性费用。"),
            ("招待费", "业务招待和客户接待费用。包括商务用餐、客户接待、会议餐费等合理的业务招待支出。"),
            ("广告费", "市场推广和广告宣传费用。包括线上广告投放、宣传物料制作、展会参展、品牌推广等营销支出。"),
            ("运输费", "货物运输和物流费用。包括快递费、物流配送费、货运费、仓储费等运输相关支出。"),
            ("手续费", "各类金融和行政手续费用。包括银行手续费、汇款费、认证费、审批费、服务费等。"),
            ("营业外支出", "非日常经营活动产生的支出。包括捐赠支出、罚款支出、资产处置损失等特殊项目支出。"),
            ("工资", "员工薪酬支出。包括基本工资、奖金、津贴、补贴等人力成本相关的工资性支出。"),
            ("租赁费", "租赁相关费用。包括办公场所租金、设备租赁费、车辆租赁费等各类租赁合同项下的支出。"),
            ("其他应收", "其他类型的应收款项和杂项支出。包括临时性支出、预付款项、保证金等暂时无法归类的费用。")
        ]
        
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        
        for name, description in categories_data:
            existing_category = Category.query.filter_by(name=name).first()
            
            if existing_category:
                # 更新描述
                if existing_category.description != description:
                    existing_category.description = description
                    updated_count += 1
                else:
                    skipped_count += 1
            else:
                # 创建新类型
                new_category = Category(name=name, description=description, created_by=session['user_id'])
                db.session.add(new_category)
                imported_count += 1
        
        db.session.commit()
        
        return jsonify({
            'message': '报销类型导入成功',
            'imported': imported_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'total': len(categories_data)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'导入失败：{str(e)}'}), 500

# 数据库清除功能
@app.route('/api/admin/clear-database', methods=['POST'])
def clear_database():
    """清除数据库所有数据（仅超级管理员）"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '权限不足'}), 403
    
    try:
        data = request.get_json()
        confirmation_text = data.get('confirmation', '')
        admin_password = data.get('admin_password', '')
        
        # 安全验证：必须输入确认文本
        if confirmation_text != 'CLEAR ALL DATA':
            return jsonify({'success': False, 'message': '确认文本不正确'}), 400
        
        # 验证管理员密码
        admin_user = User.query.get(session['user_id'])
        if not admin_user or admin_user.password != admin_password:
            return jsonify({'success': False, 'message': '管理员密码错误'}), 401
        
        # 获取当前数据统计
        stats_before = get_database_stats()
        
        # 按顺序清除数据（考虑外键约束）
        clear_results = []
        
        # 1. 清除通知（依赖用户和报销）
        notification_count = Notification.query.count()
        Notification.query.delete()
        clear_results.append(f"通知记录: {notification_count} 条")
        
        # 2. 清除报销附件（依赖报销）
        file_count = ExpenseFile.query.count()
        ExpenseFile.query.delete()
        clear_results.append(f"附件记录: {file_count} 条")
        
        # 3. 清除报销记录（依赖用户）
        expense_count = Expense.query.count()
        Expense.query.delete()
        clear_results.append(f"报销记录: {expense_count} 条")
        
        # 4. 清除货币（依赖用户创建者）
        currency_count = Currency.query.count()
        Currency.query.delete()
        clear_results.append(f"货币记录: {currency_count} 条")
        
        # 5. 清除分类（依赖用户创建者）
        category_count = Category.query.count()
        Category.query.delete()
        clear_results.append(f"分类记录: {category_count} 条")
        
        # 6. 清除用户（最后清除，其他表都依赖用户）
        user_count = User.query.count()
        User.query.delete()
        clear_results.append(f"用户记录: {user_count} 条")
        
        # 提交清除操作
        db.session.commit()
        
        # 清除上传文件
        file_cleanup_result = cleanup_uploaded_files()
        clear_results.append(f"上传文件: {file_cleanup_result}")
        
        return jsonify({
            'success': True,
            'message': '数据库已完全清除',
            'stats_before': stats_before,
            'clear_results': clear_results,
            'timestamp': datetime.utcnow().isoformat()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'清除失败：{str(e)}'}), 500

@app.route('/api/admin/database-stats', methods=['GET'])
def get_database_stats_api():
    """获取数据库统计信息"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'error': '权限不足'}), 403
    
    try:
        stats = get_database_stats()
        return jsonify({'success': True, 'stats': stats})
    except Exception as e:
        return jsonify({'success': False, 'message': f'获取统计失败：{str(e)}'}), 500

def get_database_stats():
    """获取数据库各表的记录统计"""
    stats = {
        'users': User.query.count(),
        'expenses': Expense.query.count(),
        'expense_files': ExpenseFile.query.count(),
        'notifications': Notification.query.count(),
        'currencies': Currency.query.count(),
        'categories': Category.query.count(),
        'total_records': 0
    }
    
    stats['total_records'] = sum(stats.values()) - stats['total_records']  # 不重复计算total_records
    
    # 获取文件统计
    try:
        upload_path = app.config['UPLOAD_FOLDER']
        if os.path.exists(upload_path):
            files = [f for f in os.listdir(upload_path) if os.path.isfile(os.path.join(upload_path, f))]
            stats['uploaded_files'] = len(files)
            
            # 计算文件总大小
            total_size = 0
            for filename in files:
                filepath = os.path.join(upload_path, filename)
                total_size += os.path.getsize(filepath)
            stats['total_file_size'] = total_size
        else:
            stats['uploaded_files'] = 0
            stats['total_file_size'] = 0
    except Exception:
        stats['uploaded_files'] = 0
        stats['total_file_size'] = 0
    
    return stats

def cleanup_uploaded_files():
    """清理上传的文件"""
    try:
        upload_path = app.config['UPLOAD_FOLDER']
        if not os.path.exists(upload_path):
            return "上传目录不存在"
        
        files = [f for f in os.listdir(upload_path) if os.path.isfile(os.path.join(upload_path, f))]
        deleted_count = 0
        
        for filename in files:
            try:
                file_path = os.path.join(upload_path, filename)
                os.remove(file_path)
                deleted_count += 1
            except Exception as e:
                print(f"删除文件失败 {filename}: {e}")
        
        return f"{deleted_count} 个文件已删除"
        
    except Exception as e:
        return f"文件清理失败: {str(e)}"

@app.route('/api/admin/reset-to-defaults', methods=['POST'])
def reset_to_defaults():
    """重置数据库到默认状态（清除数据后重新创建默认数据）"""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '权限不足'}), 403
    
    try:
        data = request.get_json()
        confirmation_text = data.get('confirmation', '')
        admin_password = data.get('admin_password', '')
        
        # 安全验证
        if confirmation_text != 'RESET TO DEFAULTS':
            return jsonify({'success': False, 'message': '确认文本不正确'}), 400
        
        # 验证管理员密码
        admin_user = User.query.get(session['user_id'])
        if not admin_user or admin_user.password != admin_password:
            return jsonify({'success': False, 'message': '管理员密码错误'}), 401
        
        # 清除所有数据
        clear_results = []
        
        # 清除顺序（考虑外键约束）
        Notification.query.delete()
        ExpenseFile.query.delete()
        Expense.query.delete()
        Currency.query.delete()
        Category.query.delete()
        User.query.delete()
        
        db.session.commit()
        
        # 清理上传文件
        cleanup_uploaded_files()
        
        # 重新创建默认数据
        create_default_data()
        
        return jsonify({
            'success': True,
            'message': '数据库已重置为默认状态',
            'timestamp': datetime.utcnow().isoformat()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'重置失败：{str(e)}'}), 500

def create_default_data():
    """创建默认数据"""
    # 创建默认管理员
    admin_user = User(
        username='管理员',
        email='admin@company.com',
        password='admin123',
        role='admin'
    )
    db.session.add(admin_user)
    db.session.flush()  # 获取admin_user.id
    
    # 创建默认员工
    employee_user = User(
        username='张三',
        email='user@company.com',
        password='user123',
        role='employee'
    )
    db.session.add(employee_user)
    db.session.flush()
    
    # 创建默认货币
    default_currencies = [
        Currency(code='USD', name='美元', symbol='$', exchange_rate=1.0000, created_by=admin_user.id),
        Currency(code='CNY', name='人民币', symbol='¥', exchange_rate=7.2500, created_by=admin_user.id),
        Currency(code='EUR', name='欧元', symbol='€', exchange_rate=0.9091, created_by=admin_user.id),
        Currency(code='GBP', name='英镑', symbol='£', exchange_rate=0.8000, created_by=admin_user.id),
        Currency(code='JPY', name='日元', symbol='¥', exchange_rate=149.25, created_by=admin_user.id),
        Currency(code='KRW', name='韩元', symbol='₩', exchange_rate=1428.57, created_by=admin_user.id),
        Currency(code='HKD', name='港元', symbol='HK$', exchange_rate=7.8000, created_by=admin_user.id),
        Currency(code='SGD', name='新加坡元', symbol='S$', exchange_rate=1.3514, created_by=admin_user.id),
        Currency(code='AUD', name='澳元', symbol='A$', exchange_rate=1.4925, created_by=admin_user.id),
        Currency(code='CAD', name='加元', symbol='C$', exchange_rate=1.3514, created_by=admin_user.id),
    ]
    
    for currency in default_currencies:
        db.session.add(currency)
    
    # 创建默认分类
    default_categories = [
        Category(name='餐饮费', description='工作餐、客户招待等餐饮费用', created_by=admin_user.id),
        Category(name='交通费', description='出差交通、打车、公交等交通费用', created_by=admin_user.id),
        Category(name='办公费', description='办公用品、文具、设备等办公费用', created_by=admin_user.id),
        Category(name='差旅费', description='出差住宿、机票等差旅费用', created_by=admin_user.id),
        Category(name='通讯费', description='电话、网络、邮寄等通讯费用', created_by=admin_user.id),
        Category(name='其他', description='其他类型的报销费用', created_by=admin_user.id),
    ]
    
    for category in default_categories:
        db.session.add(category)
    
    db.session.commit()

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # 检查并创建默认数据
        if User.query.count() == 0:
            # 创建默认管理员
            admin_user = User(
                username='管理员',
                email='admin@company.com',
                password='admin123',
                role='admin'
            )
            db.session.add(admin_user)
            db.session.commit()
            
            # 创建默认货币
            default_currencies = [
                Currency(code='USD', name='美元', symbol='$', exchange_rate=1.0000, created_by=admin_user.id),
                Currency(code='CNY', name='人民币', symbol='¥', exchange_rate=0.1379, created_by=admin_user.id),
                Currency(code='EUR', name='欧元', symbol='€', exchange_rate=1.1000, created_by=admin_user.id),
                Currency(code='GBP', name='英镭', symbol='£', exchange_rate=1.2500, created_by=admin_user.id),
                Currency(code='JPY', name='日元', symbol='¥', exchange_rate=0.0067, created_by=admin_user.id),
            ]
            
            for currency in default_currencies:
                db.session.add(currency)
            
            # 创建默认分类
            default_categories = [
                Category(name='餐饮费', description='工作餐、客户招待等餐饮费用', created_by=admin_user.id),
                Category(name='交通费', description='出差交通、打车、公交等交通费用', created_by=admin_user.id),
                Category(name='办公费', description='办公用品、文具、设备等办公费用', created_by=admin_user.id),
                Category(name='差旅费', description='出差住宿、机票等差旅费用', created_by=admin_user.id),
                Category(name='通讯费', description='电话、网络、邮寄等通讯费用', created_by=admin_user.id),
                Category(name='其他', description='其他类型的报销费用', created_by=admin_user.id),
            ]
            
            for category in default_categories:
                db.session.add(category)
            
            db.session.commit()
            print('默认数据创建完成')
    
    app.run(debug=True)